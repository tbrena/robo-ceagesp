# -*- coding: utf-8 -*-
"""
Robo CEAGESP - Cotacoes de Pescado no Atacado
==============================================
Coleta diariamente os precos de atacado dos pescados de interesse
em https://ceagesp.gov.br/cotacoes/ e acumula o historico.

Uso:
    python robo_ceagesp.py                  # coleta o(s) boletim(ns) ainda nao coletado(s)
    python robo_ceagesp.py --data 03/08/2026  # coleta uma data especifica
    python robo_ceagesp.py --todas          # recoleta todas as datas disponiveis no site
    python robo_ceagesp.py --tudo           # coleta todos os produtos da categoria (nao so os da lista)

Saidas (pasta ./dados):
    historico.csv                 - historico acumulado (nao duplica)
    boletim_AAAA-MM-DD.csv        - cada boletim coletado
    cotacoes_pescado.xlsx         - planilha com Historico + Ultimo boletim
"""

import argparse
import csv
import json
import logging
import os
import re
import sys
import time
import unicodedata
from datetime import datetime

import requests
import urllib3

# --------------------------------------------------------------------------
# CONFIGURACAO
# --------------------------------------------------------------------------

URL = "https://ceagesp.gov.br/cotacoes/"
CATEGORIA = "PESCADOS"

# Produtos de interesse. Comparacao ignora acentos, maiusculas e espacos extras.
PRODUTOS = [
    "FILE DE TILAPIA",
    "PANGASIUS",
    "PINTADO",
    "TAMBAQUI",
    "TILAPIA",
    "TRUTA",
]

BASE = os.path.dirname(os.path.abspath(__file__))
DIR_DADOS = os.path.join(BASE, "dados")
DIR_LOGS = os.path.join(BASE, "logs")
CSV_HISTORICO = os.path.join(DIR_DADOS, "historico.csv")
XLSX = os.path.join(DIR_DADOS, "cotacoes_pescado.xlsx")

COLUNAS = [
    "data", "categoria", "produto", "classificacao", "unidade_peso",
    "menor", "comum", "maior", "quilo", "coletado_em",
]

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36")

TENTATIVAS = 3
ESPERA_ENTRE_TENTATIVAS = 10  # segundos


# --------------------------------------------------------------------------
# INFRAESTRUTURA
# --------------------------------------------------------------------------

def configurar_log():
    os.makedirs(DIR_LOGS, exist_ok=True)
    log = logging.getLogger("robo")
    log.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%d/%m/%Y %H:%M:%S")

    arquivo = logging.FileHandler(os.path.join(DIR_LOGS, "robo.log"), encoding="utf-8")
    arquivo.setFormatter(fmt)
    log.addHandler(arquivo)

    console = logging.StreamHandler(sys.stdout)
    console.setFormatter(fmt)
    log.addHandler(console)
    return log


log = configurar_log()


def normalizar(texto):
    """Remove acentos, colapsa espacos e passa para maiuscula."""
    texto = unicodedata.normalize("NFKD", texto or "")
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", texto).strip().upper()


ALVOS = {normalizar(p) for p in PRODUTOS}


def para_numero(txt):
    """'42,52' -> 42.52 ; devolve None quando nao houver numero."""
    txt = (txt or "").strip().replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return None


def requisitar(sessao, metodo, **kwargs):
    """GET/POST com retentativa. Se o certificado do site falhar (ja aconteceu
    com o ceagesp.gov.br), refaz sem verificacao TLS e registra aviso."""
    ultimo_erro = None
    for tentativa in range(1, TENTATIVAS + 1):
        for verificar in (True, False):
            try:
                if not verificar:
                    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
                resp = sessao.request(metodo, URL, timeout=60, verify=verificar, **kwargs)
                resp.raise_for_status()
                if not verificar:
                    log.warning("Certificado TLS do site invalido - requisicao feita SEM "
                                "verificacao de certificado.")
                return resp
            except requests.exceptions.SSLError as e:
                ultimo_erro = e
                continue  # tenta de novo sem verificar
            except Exception as e:
                ultimo_erro = e
                break     # erro nao-TLS: nao adianta tentar sem verificar
        if tentativa < TENTATIVAS:
            log.warning("Falha na tentativa %d/%d (%s). Nova tentativa em %ds.",
                        tentativa, TENTATIVAS, type(ultimo_erro).__name__,
                        ESPERA_ENTRE_TENTATIVAS)
            time.sleep(ESPERA_ENTRE_TENTATIVAS)
    raise RuntimeError("Nao foi possivel acessar %s: %s" % (URL, ultimo_erro))


# --------------------------------------------------------------------------
# COLETA
# --------------------------------------------------------------------------

def datas_disponiveis(sessao):
    """Le a variavel javascript 'Grupos' da pagina, que traz as datas com
    boletim publicado para cada categoria."""
    html = requisitar(sessao, "GET").text
    m = re.search(r"var\s+Grupos\s*=\s*(\{.*?\})\s*;", html, re.S)
    if not m:
        raise RuntimeError("Nao encontrei a lista de datas na pagina (layout do site mudou?).")

    grupos = json.loads(m.group(1))
    chave = next((k for k in grupos if normalizar(k) == normalizar(CATEGORIA)), None)
    if chave is None:
        raise RuntimeError("Categoria %s nao existe no site. Disponiveis: %s"
                           % (CATEGORIA, list(grupos)))

    datas = grupos.get(chave) or []
    return sorted(datas, key=lambda d: datetime.strptime(d, "%d/%m/%Y"))


def coletar_boletim(sessao, data, todos_produtos=False):
    """Consulta uma data e devolve a lista de linhas do boletim."""
    resp = requisitar(sessao, "POST", data={"cot_grupo": CATEGORIA, "cot_data": data})
    html = resp.text

    tabela = re.search(r'<table[^>]*class="[^"]*contacao_lista[^"]*".*?</table>', html, re.S)
    if not tabela:
        log.warning("Boletim de %s: nenhuma tabela retornada.", data)
        return []

    agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    linhas = []
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", tabela.group(0), re.S):
        celulas = [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", td)).strip()
                   for td in re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S)]
        if len(celulas) != 7 or normalizar(celulas[0]) in ("PRODUTO", ""):
            continue  # cabecalho ou linha de titulo

        produto = re.sub(r"\s+", " ", celulas[0]).strip()
        if not todos_produtos and normalizar(produto) not in ALVOS:
            continue

        linhas.append({
            "data": data,
            "categoria": CATEGORIA,
            "produto": produto,
            "classificacao": celulas[1],
            "unidade_peso": celulas[2],
            "menor": para_numero(celulas[3]),
            "comum": para_numero(celulas[4]),
            "maior": para_numero(celulas[5]),
            "quilo": para_numero(celulas[6]),
            "coletado_em": agora,
        })

    if not todos_produtos:
        achados = {normalizar(l["produto"]) for l in linhas}
        for faltante in sorted(ALVOS - achados):
            log.warning("Boletim de %s: produto '%s' nao foi cotado nesta data.", data, faltante)

    return linhas


# --------------------------------------------------------------------------
# ARMAZENAMENTO
# --------------------------------------------------------------------------

def ler_historico():
    if not os.path.exists(CSV_HISTORICO):
        return []
    with open(CSV_HISTORICO, encoding="utf-8-sig", newline="") as f:
        registros = list(csv.DictReader(f, delimiter=";"))
    for r in registros:
        for c in ("menor", "comum", "maior", "quilo"):
            r[c] = para_numero(r.get(c))
    return registros


def gravar_csv(caminho, registros):
    with open(caminho, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUNAS, delimiter=";")
        w.writeheader()
        for r in registros:
            w.writerow({c: ("" if r.get(c) is None else
                            (str(r[c]).replace(".", ",") if c in ("menor", "comum", "maior", "quilo")
                             else r[c]))
                        for c in COLUNAS})


def ordenar(registros):
    return sorted(registros, key=lambda r: (datetime.strptime(r["data"], "%d/%m/%Y"),
                                            normalizar(r["produto"]),
                                            r.get("classificacao") or ""))


def gerar_excel(historico):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    if not historico:
        return

    ultima = max(historico, key=lambda r: datetime.strptime(r["data"], "%d/%m/%Y"))["data"]
    ultimo_boletim = [r for r in historico if r["data"] == ultima]

    wb = Workbook()
    cabecalho_fill = PatternFill("solid", fgColor="1F6F43")
    cabecalho_fonte = Font(bold=True, color="FFFFFF")

    for nome, dados in (("Historico", historico), ("Ultimo boletim", ultimo_boletim)):
        ws = wb.active if nome == "Historico" else wb.create_sheet()
        ws.title = nome
        ws.append([c.replace("_", " ").capitalize() for c in COLUNAS])
        for celula in ws[1]:
            celula.fill = cabecalho_fill
            celula.font = cabecalho_fonte
            celula.alignment = Alignment(horizontal="center")

        for r in dados:
            ws.append([r.get(c) for c in COLUNAS])

        for i, col in enumerate(COLUNAS, start=1):
            letra = get_column_letter(i)
            if col in ("menor", "comum", "maior", "quilo"):
                for celula in ws[letra][1:]:
                    celula.number_format = 'R$ #,##0.00'
            largura = max([len(str(r.get(col) or "")) for r in dados] + [len(col)]) + 3
            ws.column_dimensions[letra].width = min(largura, 28)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(XLSX)


# --------------------------------------------------------------------------
# PRINCIPAL
# --------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Robo de cotacoes de pescado da CEAGESP")
    ap.add_argument("--data", help="coleta uma data especifica (DD/MM/AAAA)")
    ap.add_argument("--todas", action="store_true",
                    help="recoleta todas as datas disponiveis no site")
    ap.add_argument("--tudo", action="store_true",
                    help="coleta todos os produtos da categoria, nao so a lista configurada")
    args = ap.parse_args()

    os.makedirs(DIR_DADOS, exist_ok=True)
    log.info("=" * 62)
    log.info("Iniciando coleta - categoria %s", CATEGORIA)

    sessao = requests.Session()
    sessao.headers.update({"User-Agent": UA})

    disponiveis = datas_disponiveis(sessao)
    log.info("Datas publicadas no site: %s", ", ".join(disponiveis) or "(nenhuma)")

    historico = ler_historico()
    ja_coletadas = {r["data"] for r in historico}

    if args.data:
        alvo = [args.data]
        if args.data not in disponiveis:
            log.warning("A data %s nao consta na lista do site; consultando assim mesmo.", args.data)
    elif args.todas:
        alvo = disponiveis
    else:
        alvo = [d for d in disponiveis if d not in ja_coletadas]

    if not alvo:
        log.info("Nenhum boletim novo. Historico ja esta em dia (%d registros).", len(historico))
        from gerar_pagina import SAIDA, gerar as gerar_pagina
        from gerar_api import DIR_API, gerar as gerar_api
        if not os.path.exists(SAIDA):
            log.info("Pagina de consulta ausente - gerando: %s", gerar_pagina(historico))
        if not os.path.exists(DIR_API):
            log.info("API ausente - gerando: %s", gerar_api(historico))
        return 0

    novos = []
    for data in alvo:
        log.info("Consultando boletim de %s...", data)
        linhas = coletar_boletim(sessao, data, todos_produtos=args.tudo)
        if not linhas:
            log.warning("Boletim de %s nao trouxe nenhum produto de interesse.", data)
            continue
        gravar_csv(os.path.join(DIR_DADOS, "boletim_%s.csv"
                                % datetime.strptime(data, "%d/%m/%Y").strftime("%Y-%m-%d")),
                   ordenar(linhas))
        for l in linhas:
            log.info("   %-18s %-3s  menor %-7s comum %-7s maior %-7s",
                     l["produto"], l["classificacao"], l["menor"], l["comum"], l["maior"])
        novos.extend(linhas)
        time.sleep(2)  # gentileza com o servidor

    if not novos:
        log.info("Nada novo gravado.")
        return 0

    # Junta com o historico substituindo registros repetidos (data+produto+classificacao)
    chave = lambda r: (r["data"], normalizar(r["produto"]), r.get("classificacao") or "")
    consolidado = {chave(r): r for r in historico}
    consolidado.update({chave(r): r for r in novos})
    historico = ordenar(list(consolidado.values()))

    gravar_csv(CSV_HISTORICO, historico)
    gerar_excel(historico)

    from gerar_pagina import gerar as gerar_pagina
    log.info("Pagina de consulta gerada: %s", gerar_pagina(historico))
    from gerar_api import gerar as gerar_api
    log.info("Endpoints da API gerados: %s", gerar_api(historico))

    log.info("Concluido: %d registros novos/atualizados | historico com %d linhas.",
             len(novos), len(historico))
    log.info("Arquivos: %s", DIR_DADOS)
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception as erro:
        log.exception("ERRO na execucao: %s", erro)
        sys.exit(1)
