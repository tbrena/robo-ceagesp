# -*- coding: utf-8 -*-
"""
Importa historico antigo a partir de uma planilha Excel para dentro do
dados/historico.csv, sem duplicar e sem sobrescrever o que o robo coletou.

A planilha precisa ter uma linha de cabecalho com as colunas
Data / Produto / Clas. / Uni/Peso / Menor / Comum / Maior (a ordem nao importa,
e variacoes de acento e maiuscula sao aceitas).

Uso:
    python importar_excel.py "C:\\caminho\\Precos Pescado CEAGESP.xlsx"
    python importar_excel.py planilha.xlsx --aba PESCADO
    python importar_excel.py planilha.xlsx --todos      # importa todos os produtos
    python importar_excel.py planilha.xlsx --simular    # so mostra o que faria

Em caso de conflito (mesma data + produto + classificacao), o registro ja
existente prevalece: o robo le direto do site da CEAGESP, entao ele e a
referencia. As divergencias sao listadas no final para conferencia.
"""

import argparse
import os
import sys
from datetime import datetime

from openpyxl import load_workbook

from robo_ceagesp import (ALVOS, CSV_HISTORICO, gerar_excel, gravar_csv,
                          ler_historico, log, normalizar, ordenar)

# Cabecalhos aceitos -> nome interno da coluna
CABECALHOS = {
    "DATA": "data",
    "PRODUTO": "produto",
    "CLAS.": "classificacao", "CLAS": "classificacao", "CLASSIFICACAO": "classificacao",
    "UNI/PESO": "unidade_peso", "UNI PESO": "unidade_peso", "UNIDADE": "unidade_peso",
    "MENOR": "menor", "COMUM": "comum", "MAIOR": "maior",
    "KG": "quilo", "QUILO": "quilo",
}
OBRIGATORIAS = {"data", "produto", "menor", "comum", "maior"}


def achar_cabecalho(ws, limite=15):
    """Localiza a linha de cabecalho e mapeia coluna -> nome interno."""
    for linha in ws.iter_rows(min_row=1, max_row=limite):
        mapa = {}
        for celula in linha:
            nome = CABECALHOS.get(normalizar(celula.value).rstrip(":"))
            if nome:
                mapa[celula.column - 1] = nome
        if OBRIGATORIAS <= set(mapa.values()):
            return linha[0].row, mapa
    raise RuntimeError(
        "Nao encontrei a linha de cabecalho nas primeiras %d linhas da aba '%s'. "
        "Esperado algo como: Data | Produto | Clas. | Uni/Peso | Menor | Comum | Maior"
        % (limite, ws.title))


def numero(v):
    if isinstance(v, (int, float)):
        return float(v)
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(".", "").replace(",", "."))
    except ValueError:
        return None


def ler_planilha(caminho, aba=None, todos=False):
    wb = load_workbook(caminho, data_only=True)
    ws = wb[aba] if aba else wb[wb.sheetnames[0]]
    inicio, colunas = achar_cabecalho(ws)
    log.info("Aba '%s': cabecalho na linha %d, colunas %s",
             ws.title, inicio, sorted(set(colunas.values())))

    origem = "%s (%s)" % (os.path.basename(caminho), ws.title)
    carimbo = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    registros, ignoradas, fora = [], 0, 0
    for linha in ws.iter_rows(min_row=inicio + 1, values_only=True):
        bruto = {nome: (linha[i] if i < len(linha) else None)
                 for i, nome in colunas.items()}

        data = bruto.get("data")
        if isinstance(data, datetime):
            data = data.strftime("%d/%m/%Y")
        else:
            try:
                data = datetime.strptime(str(data).strip()[:10], "%d/%m/%Y").strftime("%d/%m/%Y")
            except (ValueError, TypeError):
                ignoradas += 1
                continue

        produto = " ".join(str(bruto.get("produto") or "").split())
        if not produto:
            ignoradas += 1
            continue
        if not todos and normalizar(produto) not in ALVOS:
            fora += 1
            continue

        registros.append({
            "data": data,
            "categoria": "PESCADOS",
            "produto": produto,
            "classificacao": (str(bruto.get("classificacao")).strip()
                              if bruto.get("classificacao") not in (None, "") else "-"),
            "unidade_peso": (str(bruto.get("unidade_peso")).strip()
                             if bruto.get("unidade_peso") not in (None, "") else "KG"),
            "menor": numero(bruto.get("menor")),
            "comum": numero(bruto.get("comum")),
            "maior": numero(bruto.get("maior")),
            "quilo": numero(bruto.get("quilo")) or 1.0,
            "coletado_em": "importado de %s em %s" % (origem, carimbo),
        })

    log.info("Planilha lida: %d registros aproveitados, %d de outros produtos, %d linhas ignoradas",
             len(registros), fora, ignoradas)
    return registros


def main():
    ap = argparse.ArgumentParser(description="Importa historico antigo de uma planilha Excel")
    ap.add_argument("planilha", help="caminho do arquivo .xlsx")
    ap.add_argument("--aba", help="nome da aba (padrao: a primeira)")
    ap.add_argument("--todos", action="store_true",
                    help="importa todos os produtos, nao so os monitorados")
    ap.add_argument("--simular", action="store_true",
                    help="mostra o resultado sem gravar nada")
    args = ap.parse_args()

    if not os.path.exists(args.planilha):
        log.error("Arquivo nao encontrado: %s", args.planilha)
        return 1

    log.info("=" * 62)
    log.info("Importando historico de %s", args.planilha)

    novos = ler_planilha(args.planilha, args.aba, args.todos)
    if not novos:
        log.warning("Nada a importar.")
        return 0

    historico = ler_historico()
    chave = lambda r: (r["data"], normalizar(r["produto"]), r.get("classificacao") or "-")
    atual = {chave(r): r for r in historico}

    incluidos, conflitos, divergencias = 0, 0, []
    for r in novos:
        k = chave(r)
        if k in atual:
            # O robo le direto do site: em conflito, o que ja existe prevalece.
            conflitos += 1
            antigo = atual[k]
            difs = [c for c in ("menor", "comum", "maior")
                    if antigo.get(c) is not None and r.get(c) is not None
                    and round(antigo[c], 2) != round(r[c], 2)]
            if difs:
                divergencias.append((k, {c: (r[c], antigo[c]) for c in difs}))
            continue
        atual[k] = r
        incluidos += 1

    historico = ordenar(list(atual.values()))
    datas = sorted({r["data"] for r in historico},
                   key=lambda d: datetime.strptime(d, "%d/%m/%Y"))

    log.info("Registros novos incluidos: %d", incluidos)
    log.info("Ja existentes (mantido o do robo): %d", conflitos)
    log.info("Historico passa a ter %d registros em %d boletins, de %s a %s",
             len(historico), len(datas), datas[0], datas[-1])

    if divergencias:
        log.warning("%d registro(s) da planilha divergem do que o robo coletou "
                    "(valor da planilha -> valor mantido):", len(divergencias))
        for k, difs in divergencias:
            log.warning("   %s %s [%s]: %s", k[0], k[1], k[2],
                        ", ".join("%s %.2f -> %.2f" % (c, v[0], v[1]) for c, v in difs.items()))

    if args.simular:
        log.info("Simulacao: nada foi gravado.")
        return 0

    gravar_csv(CSV_HISTORICO, historico)
    gerar_excel(historico)
    from gerar_pagina import gerar as gerar_pagina
    log.info("Pagina regerada: %s", gerar_pagina(historico))
    from gerar_api import gerar as gerar_api
    log.info("API regerada: %s", gerar_api(historico))
    log.info("Concluido.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
